#!/usr/bin/env python3
"""
Expansion Scoring Model Generator

Generates an Excel-based scoring calculator from Salesforce and Hex data.

Usage:
    python generate_scoring.py
    python generate_scoring.py --salesforce data/sf.csv --hex data/hex.csv
    python generate_scoring.py --summary
    python generate_scoring.py --export-tier 1 --format csv
"""

import argparse
import pandas as pd
from pathlib import Path
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

# Scoring weights
SCORING = {
    'license_type': {
        'Named User': 7,
        'Defined Group': 5,
        'Named User (Chargeback)': 4,
        'Sitewide': 4,
        'Right to Deploy': 3,
    },
    'health_status': {
        'Extra Green': 25,
        'Green': 20,
        'Yellow': 8,
        'Red': 3,
    },
    'free_users': [
        (50, 25), (30, 22), (20, 18), (10, 14), (5, 10), (1, 5), (0, 0)
    ],
    'activation': [
        (1.0, 18), (0.95, 16), (0.90, 14), (0.80, 10), (0.70, 6), (0, 2)
    ],
    'arr': [
        (25000, 10), (15000, 8), (10000, 6), (5000, 4), (0, 2)
    ],
    'renewal_anytime': [
        (30, 15), (60, 13), (90, 11), (120, 9), (180, 6), (999, 3)
    ],
    'renewal_dg_sw': [
        (30, 20), (60, 17), (90, 14), (120, 10), (180, 5), (999, 2)
    ],
}

TIER_THRESHOLDS = [(75, 'Tier 1'), (55, 'Tier 2'), (35, 'Tier 3'), (0, 'Below Threshold')]

# Styling
COLORS = {
    'header': '1E3A5F',
    'tier1': 'C8E6C9',
    'tier2': 'FFF9C4',
    'tier3': 'FFCCBC',
    'input': 'FFFDE7',
    'dg_sw': 'E8EAF6',
    'urgent': 'FFCDD2',
    'gray': 'F5F5F5',
}


# ============================================================================
# DATA PROCESSING
# ============================================================================

def load_salesforce_data(filepath: Path) -> pd.DataFrame:
    """Load and clean Salesforce export."""
    df = pd.read_csv(filepath, encoding='latin-1')
    
    # Parse dates
    df['Renewal Date'] = pd.to_datetime(df['Renewal Date'], format='%m/%d/%Y', errors='coerce')
    df['Days to Renewal'] = (df['Renewal Date'] - pd.Timestamp.now()).dt.days
    df['Days to Renewal'] = df['Days to Renewal'].fillna(365).astype(int)
    
    # Standardize health status
    def standardize_health(val):
        if pd.isna(val):
            return ''
        val = str(val).lower().strip()
        if 'extra' in val and 'green' in val:
            return 'Extra Green'
        elif val == 'green':
            return 'Green'
        elif val == 'yellow':
            return 'Yellow'
        elif val in ['red', 'orange']:
            return 'Red'
        return ''
    
    df['Customer Usage Health'] = df['Customer Usage Health'].apply(standardize_health)
    
    # Extract organization name
    def extract_org(name):
        name = re.sub(r'^CS FY\d{2} (Pilot )?Renewal:\s*', '', str(name))
        return name.split('//')[0].strip()
    
    df['Organization'] = df['Opportunity Name'].apply(extract_org)
    
    # Calculate true activation rate
    def calc_activation(row):
        lt = str(row.get('License Type', ''))
        filled = row.get('Seats Filled', 0) or 0
        initial = row.get('Initial Number of Seats', 0) or 0
        current = row.get('Current Number of Seats Available', 0) or 0
        
        if lt in ['Defined Group', 'Sitewide']:
            return filled / initial if initial > 0 else 0
        else:
            return filled / current if current > 0 else 0
    
    df['True Activation'] = df.apply(calc_activation, axis=1)
    
    return df


def load_hex_data(filepath) -> pd.DataFrame:
    """Load Hex free user data. Supports both old and new formats."""
    if isinstance(filepath, Path) and not filepath.exists():
        return pd.DataFrame(columns=['opportunity_name', 'total_free_users'])

    df = pd.read_csv(filepath)

    # Detect format based on columns
    if 'opportunity_name' in df.columns and 'num_free_users' in df.columns:
        # New format: aggregate by opportunity_name
        aggregated = df.groupby('opportunity_name').agg({
            'num_free_users': 'sum',
            'num_self_serve_users': 'sum',
            'num_users_with_dept_match': 'sum'
        }).reset_index()

        # Primary metric: users with department match (highest expansion potential)
        # These are free/self-serve users in the SAME department as the license
        aggregated['total_free_users'] = aggregated['num_users_with_dept_match']

        # Also keep total for reference
        aggregated['total_all_free_users'] = (
            aggregated['num_free_users'] + aggregated['num_self_serve_users']
        )
        return aggregated

    elif 'email_domain' in df.columns and 'total_expansion_users' in df.columns:
        # Old format: keep as-is for backward compatibility
        return df

    else:
        print(f"Warning: Unrecognized Hex file format. Columns: {list(df.columns)}")
        return pd.DataFrame(columns=['opportunity_name', 'total_free_users'])


def normalize_opportunity_name(name: str) -> str:
    """Normalize opportunity name for matching."""
    if pd.isna(name):
        return ''
    # Remove extra whitespace, convert to lowercase for matching
    return ' '.join(str(name).lower().split())


def merge_data(sf_df: pd.DataFrame, hex_df: pd.DataFrame) -> pd.DataFrame:
    """Merge Salesforce and Hex data by opportunity name."""
    sf_df = sf_df.copy()
    sf_df['Free Users'] = 0

    if hex_df.empty:
        return sf_df

    # Check which format we have
    if 'opportunity_name' in hex_df.columns and 'total_free_users' in hex_df.columns:
        # New format: match by opportunity name
        hex_df = hex_df.copy()
        hex_df['opp_normalized'] = hex_df['opportunity_name'].apply(normalize_opportunity_name)
        sf_df['opp_normalized'] = sf_df['Opportunity Name'].apply(normalize_opportunity_name)

        # Create lookup dict
        hex_lookup = dict(zip(hex_df['opp_normalized'], hex_df['total_free_users']))

        # Match and assign free users
        sf_df['Free Users'] = sf_df['opp_normalized'].map(hex_lookup).fillna(0).astype(int)

        # Clean up temp column
        sf_df = sf_df.drop(columns=['opp_normalized'])

        matched = (sf_df['Free Users'] > 0).sum()
        print(f"  Matched {matched}/{len(sf_df)} accounts with Hex data")

    elif 'email_domain' in hex_df.columns:
        # Old format: try to match by domain extraction from opportunity name
        hex_lookup = dict(zip(
            hex_df['email_domain'].str.lower(),
            hex_df['total_expansion_users']
        ))

        def extract_domain_and_match(opp_name):
            # Common patterns for domain extraction
            opp_lower = str(opp_name).lower()
            for domain, users in hex_lookup.items():
                # Check if domain base name appears in opportunity
                domain_base = domain.split('.')[0]
                if domain_base in opp_lower:
                    return users
            return 0

        sf_df['Free Users'] = sf_df['Opportunity Name'].apply(extract_domain_and_match).astype(int)

        matched = (sf_df['Free Users'] > 0).sum()
        print(f"  Matched {matched}/{len(sf_df)} accounts with Hex data (domain matching)")

    return sf_df


def calculate_scores(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate expansion scores for all accounts."""
    
    def get_score(value, brackets):
        for threshold, points in brackets:
            if value >= threshold:
                return points
        return 0
    
    def score_account(row):
        lt = str(row.get('License Type', ''))
        is_dg_sw = lt in ['Defined Group', 'Sitewide']
        
        # License points
        license_pts = SCORING['license_type'].get(lt, 0)
        
        # Health points
        health = str(row.get('Customer Usage Health', ''))
        health_pts = SCORING['health_status'].get(health, 0)
        
        # Activation points
        activation = row.get('True Activation', 0)
        activation_pts = get_score(activation, SCORING['activation'])
        
        # Free user points
        free_users = row.get('Free Users', 0) or 0
        free_pts = get_score(free_users, SCORING['free_users'])
        
        # ARR points
        arr = row.get('Renewal Target Amount', 0) or 0
        arr_pts = get_score(arr, SCORING['arr'])
        
        # Renewal points (different for DG/SW)
        days = row.get('Days to Renewal', 365)
        renewal_brackets = SCORING['renewal_dg_sw'] if is_dg_sw else SCORING['renewal_anytime']
        renewal_pts = 0
        for threshold, points in renewal_brackets:
            if days <= threshold:
                renewal_pts = points
                break
        
        # Total (0 if no free users identified)
        if free_users == 0:
            total = 0
            tier = 'Need Hex Data'
        else:
            total = license_pts + health_pts + activation_pts + free_pts + arr_pts + renewal_pts
            tier = 'Below Threshold'
            for threshold, tier_name in TIER_THRESHOLDS:
                if total >= threshold:
                    tier = tier_name
                    break
        
        return pd.Series({
            'License Pts': license_pts,
            'Health Pts': health_pts,
            'Activation Pts': activation_pts,
            'Free User Pts': free_pts,
            'ARR Pts': arr_pts,
            'Renewal Pts': renewal_pts,
            'Total Score': total,
            'Tier': tier,
        })
    
    scores = df.apply(score_account, axis=1)
    return pd.concat([df, scores], axis=1)


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def create_excel(df: pd.DataFrame, output_path: Path):
    """Generate the Excel scoring calculator."""
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill('solid', fgColor=COLORS['header'])
    header_font = Font(bold=True, color='FFFFFF', size=10)
    input_fill = PatternFill('solid', fgColor=COLORS['input'])
    gray_fill = PatternFill('solid', fgColor=COLORS['gray'])
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )
    
    # Split data
    anytime_df = df[df['License Type'].isin(['Named User', 'Named User (Chargeback)', 'Right to Deploy'])]
    renewal_df = df[df['License Type'].isin(['Defined Group', 'Sitewide'])]
    
    # Sheet 1: Anytime Expansion
    ws1 = wb.active
    ws1.title = "Anytime Expansion"
    _write_scoring_sheet(ws1, anytime_df, header_fill, header_font, input_fill, gray_fill, thin_border, is_dg=False)
    
    # Sheet 2: Renewal-Only Expansion
    ws2 = wb.create_sheet("Renewal-Only Expansion")
    _write_scoring_sheet(ws2, renewal_df, header_fill, header_font, input_fill, gray_fill, thin_border, is_dg=True)
    
    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    _write_summary_sheet(ws3, df, anytime_df, renewal_df, header_fill, header_font, thin_border)
    
    wb.save(output_path)
    print(f"✓ Generated: {output_path}")


def _write_scoring_sheet(ws, df, header_fill, header_font, input_fill, gray_fill, thin_border, is_dg=False):
    """Write a scoring sheet."""
    headers = ['Organization', 'License Type', 'Health', 'Activation', 'ARR', 
               'Days to Renewal', 'FREE USERS', 'Total Score', 'Tier']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row.get('Organization', ''))
        ws.cell(row=row_idx, column=2, value=row.get('License Type', ''))
        ws.cell(row=row_idx, column=3, value=row.get('Customer Usage Health', ''))
        ws.cell(row=row_idx, column=4, value=row.get('True Activation', 0))
        ws.cell(row=row_idx, column=4).number_format = '0%'
        ws.cell(row=row_idx, column=5, value=row.get('Renewal Target Amount', 0))
        ws.cell(row=row_idx, column=5).number_format = '$#,##0'
        ws.cell(row=row_idx, column=6, value=row.get('Days to Renewal', 0))
        ws.cell(row=row_idx, column=7, value=row.get('Free Users', 0))
        ws.cell(row=row_idx, column=7).fill = input_fill
        ws.cell(row=row_idx, column=8, value=row.get('Total Score', 0))
        ws.cell(row=row_idx, column=9, value=row.get('Tier', ''))
        
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = thin_border


def _write_summary_sheet(ws, df, anytime_df, renewal_df, header_fill, header_font, thin_border):
    """Write summary sheet."""
    ws['A1'] = "EXPANSION PIPELINE SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = "Total Accounts"
    ws['B3'] = len(df)
    ws['A4'] = "Anytime Expansion"
    ws['B4'] = len(anytime_df)
    ws['A5'] = "Renewal-Only"
    ws['B5'] = len(renewal_df)


# ============================================================================
# CLI
# ============================================================================

def print_summary(df: pd.DataFrame):
    """Print account summary to terminal."""
    print("\n" + "=" * 60)
    print("EXPANSION SCORING SUMMARY")
    print("=" * 60)
    
    print(f"\nTotal Accounts: {len(df)}")
    
    anytime = df[df['License Type'].isin(['Named User', 'Named User (Chargeback)', 'Right to Deploy'])]
    renewal = df[df['License Type'].isin(['Defined Group', 'Sitewide'])]
    
    print(f"\nAnytime Expansion: {len(anytime)} accounts")
    print(f"Renewal-Only: {len(renewal)} accounts")
    
    print("\nBy Tier:")
    for tier in ['Tier 1', 'Tier 2', 'Tier 3', 'Below Threshold', 'Need Hex Data']:
        count = len(df[df['Tier'] == tier])
        print(f"  {tier}: {count}")
    
    print("\nRenewal-Only Urgency:")
    for days, label in [(30, '0-30 days'), (60, '31-60 days'), (90, '61-90 days')]:
        if days == 30:
            count = len(renewal[renewal['Days to Renewal'] <= days])
        else:
            prev = days - 30
            count = len(renewal[(renewal['Days to Renewal'] > prev) & (renewal['Days to Renewal'] <= days)])
        print(f"  {label}: {count}")


def find_hex_file(specified_path: Path) -> Path:
    """Find Hex data file, checking multiple locations and patterns."""
    # If specified path exists, use it
    if specified_path.exists():
        return specified_path

    # Check for new format files in current directory and data/
    search_dirs = [Path('.'), Path('data')]
    patterns = [
        'csv_export_-_users_by_opportunity_*.csv',
        'hex_free_users.csv',
        '*users*opportunity*.csv',
        '*hex*.csv'
    ]

    for search_dir in search_dirs:
        if not search_dir.exists():
            continue
        for pattern in patterns:
            matches = list(search_dir.glob(pattern))
            if matches:
                # Return most recently modified file
                matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                return matches[0]

    return specified_path  # Return original path (will show as not found)


def main():
    parser = argparse.ArgumentParser(description='Expansion Scoring Model Generator')
    parser.add_argument('--salesforce', type=Path, default=Path('data/salesforce_export.csv'),
                        help='Path to Salesforce CSV export')
    parser.add_argument('--hex', type=Path, default=Path('data/hex_free_users.csv'),
                        help='Path to Hex free users CSV (auto-detects new format files)')
    parser.add_argument('--output', type=Path, default=Path('output/expansion-scoring.xlsx'),
                        help='Output Excel file path')
    parser.add_argument('--summary', action='store_true',
                        help='Print summary only, no Excel generation')
    parser.add_argument('--export-tier', type=int, choices=[1, 2, 3],
                        help='Export specific tier to CSV')
    parser.add_argument('--format', choices=['csv', 'json'], default='csv',
                        help='Export format for --export-tier')

    args = parser.parse_args()

    # Load data
    if not args.salesforce.exists():
        print(f"Error: Salesforce file not found: {args.salesforce}")
        print("Place your Salesforce export in data/salesforce_export.csv")
        sys.exit(1)

    print(f"Loading Salesforce data from {args.salesforce}...")
    sf_df = load_salesforce_data(args.salesforce)

    # Auto-detect Hex file
    hex_path = find_hex_file(args.hex)
    if hex_path != args.hex:
        print(f"Auto-detected Hex file: {hex_path}")
    print(f"Loading Hex data from {hex_path}...")
    hex_df = load_hex_data(hex_path)
    
    print("Merging data...")
    df = merge_data(sf_df, hex_df)
    
    print("Calculating scores...")
    df = calculate_scores(df)
    
    if args.summary:
        print_summary(df)
        return
    
    if args.export_tier:
        tier_name = f'Tier {args.export_tier}'
        tier_df = df[df['Tier'] == tier_name]
        export_path = Path(f'output/tier_{args.export_tier}.{args.format}')
        export_path.parent.mkdir(exist_ok=True)
        
        if args.format == 'csv':
            tier_df.to_csv(export_path, index=False)
        else:
            tier_df.to_json(export_path, orient='records', indent=2)
        
        print(f"✓ Exported {len(tier_df)} {tier_name} accounts to {export_path}")
        return
    
    # Generate Excel
    args.output.parent.mkdir(exist_ok=True)
    create_excel(df, args.output)
    print_summary(df)


if __name__ == '__main__':
    main()
